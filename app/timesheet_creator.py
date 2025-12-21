from decimal import Decimal
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.rich_text import CellRichText
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
import os
import re
from datetime import date, datetime, time, timedelta
from dateutil.relativedelta import relativedelta
from showinfm import show_in_file_manager
from resources import resource_path
import math
from PySide6.QtWidgets import QMessageBox


# Clean the sheet name for Excel
def clean_sheet_name(name: str) -> str:
    cleaned: str = re.sub(r'[^A-Za-z0-9]', '', name)
    return cleaned[:30]

# start_date is mm/dd/yyyy format; row range is inclusive
def create_timesheets(
        source_file_path: str,
        source_sheet_name: str,
        template_file_path: str,
        output_file_path: str,
        start_date: str,
        row_range: tuple[int, int] = (2, math.inf)
    ) -> bool:

    # Load src and template workbooks
    src_wb: Workbook = openpyxl.load_workbook(source_file_path)
    src_ws = src_wb[source_sheet_name]
    tmpl_wb: Workbook = openpyxl.load_workbook(template_file_path)
    base_sheet = tmpl_wb["Template"]

    # Get the indecies of the total hours, reg hours, and OT hours
    headers = [cell.value for cell in next(src_ws.iter_rows(min_row=1, max_row=1))]
    total_hours_col: int = headers.index("Total Hours")
    total_reg_col: int = headers.index("Total REG")
    total_ot_col: int = headers.index("Total OT")

    min_row = int(row_range[0])
    max_row = int(row_range[1]) if row_range[1] != math.inf else src_ws.max_row
    for row in src_ws.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
        name, position, location, *rest = row
        if not name: break

        # Create the new sheet for current person
        sheet_name: str = clean_sheet_name(name)
        if sheet_name in tmpl_wb.sheetnames:
            tmpl_wb.remove(tmpl_wb[sheet_name])
        new_sheet: Worksheet = tmpl_wb.copy_worksheet(base_sheet)
        new_sheet.title = sheet_name

        # Add name, and position to timesheet
        new_sheet["A1"].value = f"{name}\nPosition: {position if position else 'N/A'}"

        # add location to timesheet
        new_sheet["A16"].value = f"Location: {location}"

        # Grab and populate the clock in/out and total hours for each day
        daily_clock_data: list = rest[: (len(rest) - 3)]
        for i in range(int(len(daily_clock_data) / 3)):
            clock_in = daily_clock_data[i*3]
            clock_out = daily_clock_data[i*3+1]
            total_hours = daily_clock_data[i*3+2]

            date: date = datetime.strptime(start_date, "%m/%d/%Y").date()
            date = date + relativedelta(days=+i)

            new_sheet[f"B{i+2}"].value = clock_in
            new_sheet[f"C{i+2}"].value = clock_out
            new_sheet[f"D{i+2}"].value = total_hours
            
            new_sheet[f"A{i+2}"].value = date.strftime("%m/%d/%Y")
        
        # Grab and populate the total, reg, and OT hours
        total_hours = row[total_hours_col]
        total_reg = row[total_reg_col]
        total_ot = row[total_ot_col]
        new_sheet["D16"].value = f"Regular: {round(float(total_reg), 2)}"
        new_sheet["E16"].value = f"OverTime: {round(float(total_ot), 2)}"
        new_sheet["F16"].value = f"Total Hours: {round(float(total_hours), 2)}"
    
    tmpl_wb.remove(base_sheet)

    try:
        tmpl_wb.save(output_file_path)
        show_in_file_manager(output_file_path)
        
        return True
    except Exception as e:
        message_body = f"An error occurred while saving the timesheets. " \
                        f"Please ensure the file is not open in another program and try again.\n\nError details: {e}"
        QMessageBox.critical(
            None,
            "Error Saving Timesheets",
            message_body
        )

        return False